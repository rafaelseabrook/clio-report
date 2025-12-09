# Generate_TrafficLightReport.py
# Correct Net Trust math (ID-joins) + per-user Cycle Hours columns + SharePoint upload

import os
import re
import time
import requests
import pandas as pd
from datetime import datetime
from urllib.parse import quote
from flask import Flask, request
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, NamedStyle, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import msal
from typing import Optional, Dict, Any, List, Tuple, Union

app = Flask(__name__)

# =========================
# Config / Environment
# =========================
API_VERSION = "4"
CLIO_BASE = os.getenv("CLIO_BASE", "https://app.clio.com").rstrip("/")
CLIO_API = f"{CLIO_BASE}/api/v{API_VERSION}"
CLIO_TOKEN_URL = f"{CLIO_BASE}/oauth/token"

CLIENT_ID = os.getenv("CLIO_CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIO_CLIENT_SECRET")
REDIRECT_URI = os.getenv("CLIO_REDIRECT_URI")

PAGE_LIMIT = 200
GLOBAL_MIN_SLEEP = float(os.getenv("CLIO_GLOBAL_MIN_SLEEP_SEC", "0.5"))

session = requests.Session()
session.headers.update({"Accept": "application/json"})

# =========================
# Auth + request helpers
# =========================
def save_tokens_env(tokens: Dict[str, Any]) -> None:
    os.environ["CLIO_ACCESS_TOKEN"] = tokens.get("access_token", "")
    if tokens.get("refresh_token"):
        os.environ["CLIO_REFRESH_TOKEN"] = tokens["refresh_token"]
    # store absolute expiry (epoch seconds)
    exp = tokens.get("expires_in", 0)
    try:
        exp_at = time.time() + float(exp)
    except Exception:
        exp_at = time.time() + 3000
    os.environ["CLIO_EXPIRES_AT"] = str(exp_at)

def load_tokens_env() -> Dict[str, Any]:
    return {
        "access_token": os.getenv("CLIO_ACCESS_TOKEN", ""),
        "refresh_token": os.getenv("CLIO_REFRESH_TOKEN", ""),
        "expires_at": float(os.getenv("CLIO_EXPIRES_AT", "0") or 0),
    }

def _get_access_token() -> str:
    toks = load_tokens_env()
    now = time.time()
    if toks["access_token"] and now < toks["expires_at"]:
        return toks["access_token"]
    return _refresh_access_token(toks["refresh_token"])

def _refresh_access_token(refresh_token: str) -> str:
    if not (CLIENT_ID and CLIENT_SECRET and refresh_token):
        raise RuntimeError("Missing CLIO credentials or refresh token.")
    resp = session.post(CLIO_TOKEN_URL, data={
        "grant_type": "refresh_token",
        "refresh_token": refresh_token,
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET
    }, timeout=60)
    if resp.status_code != 200:
        raise RuntimeError(f"Refresh failed: {resp.status_code} {resp.text[:200]}")
    tokens = resp.json() or {}
    save_tokens_env(tokens)
    tok = tokens.get("access_token", "")
    session.headers["Authorization"] = f"Bearer {tok}"
    return tok

def _ensure_auth() -> None:
    if "Authorization" not in session.headers or not session.headers.get("Authorization"):
        tok = _get_access_token()
        session.headers["Authorization"] = f"Bearer {tok}"

def _sleep_with_floor(start_ts: float, retry_after: Optional[Union[int, float]] = None) -> None:
    elapsed = time.time() - start_ts
    base = max(0, GLOBAL_MIN_SLEEP - elapsed)
    wait = max(base, float(retry_after or 0))
    if wait > 0:
        time.sleep(wait)

def _request(method: str, url: str, **kwargs) -> requests.Response:
    _ensure_auth()
    max_tries = kwargs.pop("max_tries", 7)
    backoff = 1
    for _ in range(max_tries):
        t0 = time.time()
        resp = session.request(method, url, timeout=60, **kwargs)
        if resp.status_code == 401:
            _refresh_access_token(load_tokens_env()["refresh_token"])
            _sleep_with_floor(t0, 1)
            continue
        if resp.status_code == 429:
            ra = 30
            if resp.headers.get("Retry-After"):
                try:
                    ra = int(resp.headers["Retry-After"])
                except Exception:
                    ra = 30
            else:
                try:
                    msg = (resp.json() or {}).get("error", {}).get("message", "")
                    m = re.search(r"Retry in (\d+)", msg)
                    if m:
                        ra = int(m.group(1))
                except Exception:
                    pass
            _sleep_with_floor(t0, ra)
            continue
        if 500 <= resp.status_code < 600:
            _sleep_with_floor(t0, backoff)
            backoff = min(backoff * 2, 60)
            continue
        _sleep_with_floor(t0)
        return resp
    return resp

def paginate(url: str, params: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    params = dict(params or {})
    params.setdefault("limit", PAGE_LIMIT)
    params.setdefault("order", "id(asc)")
    all_rows: List[Dict[str, Any]] = []
    next_url, next_params = url, params
    while True:
        resp = _request("GET", next_url, params=next_params)
        print(f"GET {next_url} params={next_params} -> {resp.status_code}")
        if resp.status_code != 200:
            print(f"Failed page: {resp.status_code} {resp.text[:200]}")
            break
        body = resp.json() or {}
        rows = body.get("data", []) or []
        all_rows.extend([r for r in rows if isinstance(r, dict)])
        paging = (body.get("meta") or {}).get("paging") or {}
        if paging.get("next"):
            next_url, next_params = paging["next"], None
            continue
        break
    return all_rows

# =========================
# Data fetchers (ID-aware)
# =========================
def fetch_custom_fields_meta() -> Dict[str, Dict[str, Any]]:
    rows = paginate(f"{CLIO_API}/custom_fields.json",
                    {"fields": "id,name,field_type,picklist_options"})
    out: Dict[str, Dict[str, Any]] = {}
    for f in rows:
        name = f.get("name")
        if not name:
            continue
        opts = {str(o["id"]): o["option"] for o in f.get("picklist_options", []) if o.get("id")}
        out[name] = {"id": f.get("id"), "type": f.get("field_type"), "options": opts}
    return out

def fetch_open_matters_with_cf() -> List[Dict[str, Any]]:
    fields = ("id,display_number,number,client{id,name},responsible_attorney{name},"
              "matter_stage{name},account_balances{balance},"
              "custom_field_values{id,field_name,field_type,value,picklist_option}")
    return paginate(f"{CLIO_API}/matters.json", {"status": "open,pending", "fields": fields})

def fetch_outstanding_client_balances() -> List[Dict[str, Any]]:
    return paginate(f"{CLIO_API}/outstanding_client_balances.json",
                    {"fields": "contact{id,name},total_outstanding_balance"})

def fetch_billable_matters() -> List[Dict[str, Any]]:
    return paginate(f"{CLIO_API}/billable_matters.json",
                    {"fields": "id,display_number,client{id,name},unbilled_amount,unbilled_hours"})

def fetch_billable_hours(start_iso: str, end_iso: str) -> Dict[str, Any]:
    """
    Returns dict keyed by display_number:
    {
      "<Matter Number>": {
          "total_hours": float,
          "user_hours": { "User A": h, ... }
      }
    }
    """
    url = f"{CLIO_API}/activities"
    params = {
        "start_date": start_iso,
        "end_date": end_iso,
        "status": "billable",
        "order": "date(desc)",
        "limit": 50,
        "fields": "id,rounded_quantity,date,matter{id,display_number},user{name},type"
    }
    headers = {"Authorization": f"Bearer {_get_access_token()}", "Accept": "application/json"}
    matter_totals: Dict[str, Any] = {}
    offset = 0
    while True:
        t0 = time.time()
        q = dict(params)
        q["offset"] = offset
        resp = requests.get(url, headers=headers, params=q, timeout=60)
        _sleep_with_floor(t0)
        if resp.status_code != 200:
            print(f"Activities error: {resp.status_code} {resp.text[:200]}")
            return {}
        body = resp.json() or {}
        rows = body.get("data", []) or []
        if not rows:
            break
        for e in rows:
            if e.get("type") != "TimeEntry":
                continue
            dn = (e.get("matter") or {}).get("display_number")
            if not dn:
                continue
            try:
                hrs = float(e.get("rounded_quantity", 0) or 0) / 3600.0
            except Exception:
                hrs = 0.0
            user = (e.get("user") or {}).get("name", "Unknown User")
            bucket = matter_totals.setdefault(dn, {"total_hours": 0.0, "user_hours": {}})
            bucket["total_hours"] += hrs
            bucket["user_hours"][user] = bucket["user_hours"].get(user, 0.0) + hrs
        if len(rows) < params["limit"]:
            break
        offset += params["limit"]
    return matter_totals

# =========================
# Transform helpers
# =========================
CF_OUTPUT_FIELDS = [
    "CR ID","Main Paralegal","Supporting Attorney","Supporting Paralegal",
    "Initial Client Goals","Initial Strategy","Has strategy changed Describe","Detailed List of Issues in the Case",
    "Strategy to Resolve Issues in the Case","Current action Items","Client Notes",
    "All Hearings on Calendar","All Deadlines on Calendar","DV situation description (DVRO: Attorney Action Required? Please Describe)",
    "Child Custody and Child Visitation (Timeshare): Temporary Orders Needed? State Current Order",
    "CS Add ons Extracurricular. Child Support: Temporary Orders Needed? State Current Order",
    "Spousal Support: Temporary Orders Needed? State Current Order",
    "Preliminary Declaration of Disclosures: Status of Client's PDDs and OPPs PDDs","Formal Discovery: Outline Discovery Strategy",
    "Parentage / Dissolution of Marriage / Legal Separation: Please Identify",
    "Judgment: Has a Judgement Been Entered? Please Specify","collection efforts"
]

def _resolve_cf(cf: Dict[str, Any], meta_by_name: Dict[str, Dict[str, Any]]) -> str:
    ftype = cf.get("field_type")
    if ftype == "picklist":
        opt = (cf.get("picklist_option") or {}).get("option")
        if opt:
            return opt
        raw = cf.get("value")
        options = ((meta_by_name.get(cf.get("field_name")) or {}).get("options")) or {}
        return options.get(str(raw), raw) or ""
    return str(cf.get("value") or "")

def build_base_frames() -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    cf_meta = fetch_custom_fields_meta()
    matters = fetch_open_matters_with_cf()
    ocb = fetch_outstanding_client_balances()
    bill = fetch_billable_matters()

    # Matters (per matter)
    m_rows: List[Dict[str, Any]] = []
    for m in matters:
        trust = 0.0
        for b in (m.get("account_balances") or []):
            try:
                trust += float((b or {}).get("balance", 0) or 0)
            except Exception:
                pass
        cf_map: Dict[str, Any] = {}
        for cf in (m.get("custom_field_values") or []):
            name = cf.get("field_name")
            if not name:
                continue
            cf_map[name] = _resolve_cf(cf, cf_meta)

        m_rows.append({
            "Matter ID": m.get("id"),
            "Matter Number": m.get("display_number") or m.get("number") or "",
            "Client ID": (m.get("client") or {}).get("id"),
            "Client Name": (m.get("client") or {}).get("name") or "",
            "Responsible Attorney": (m.get("responsible_attorney") or {}).get("name") or "",
            "Matter Stage": (m.get("matter_stage") or {}).get("name") or "",
            "Trust Account Balance": trust,
            "_cf_map": cf_map
        })
    matter_df = pd.DataFrame(m_rows)

    # OCB (per client)
    ocb_df = pd.DataFrame([{
        "Client ID": (r.get("contact") or {}).get("id"),
        "Outstanding Balance": r.get("total_outstanding_balance", 0) or 0
    } for r in ocb])

    # Unbilled (per matter)
    bill_df = pd.DataFrame([{
        "Matter ID": bm.get("id"),
        "Matter Number": bm.get("display_number") or "",
        "Client ID": (bm.get("client") or {}).get("id"),
        "Unbilled Amount": bm.get("unbilled_amount", 0) or 0,
        "Unbilled Hours": bm.get("unbilled_hours", 0) or 0
    } for bm in bill])

    # Defensive columns
    for df, cols in [
        (matter_df, ["Matter ID","Matter Number","Client ID","Client Name","Responsible Attorney","Matter Stage","Trust Account Balance","_cf_map"]),
        (ocb_df, ["Client ID","Outstanding Balance"]),
        (bill_df, ["Matter ID","Unbilled Amount","Unbilled Hours"]),
    ]:
        for c in cols:
            if c not in df.columns:
                df[c] = 0 if ("Amount" in c or "Balance" in c or c.endswith("Hours")) else ({} if c == "_cf_map" else "")

    return matter_df, ocb_df, bill_df

def make_report_df(matter_df: pd.DataFrame,
                   ocb_df: pd.DataFrame,
                   bill_df: pd.DataFrame,
                   cycle_data: Dict[str, Any],
                   cycle_start_label: str,
                   cycle_end_label: str) -> pd.DataFrame:
    # Merge strictly: OCB by Client ID, Unbilled by Matter ID
    combined = (
        matter_df
        .merge(ocb_df, on="Client ID", how="left")
        .merge(bill_df[["Matter ID","Unbilled Amount","Unbilled Hours"]], on="Matter ID", how="left")
    )

    # Coerce numerics
    for c in ["Trust Account Balance","Outstanding Balance","Unbilled Amount","Unbilled Hours"]:
        combined[c] = pd.to_numeric(combined[c], errors="coerce").fillna(0.0)

    # Net per matter (correct)
    combined["Net Trust Account Balance"] = (
        combined["Trust Account Balance"] - combined["Outstanding Balance"] - combined["Unbilled Amount"]
    ).astype(float)

    # Cycle Hours total column
    col_total = f"Billing Cycle Hours ({cycle_start_label} - {cycle_end_label})"

    def _cycle_total(dn: str) -> float:
        b = cycle_data.get(dn, {})
        if isinstance(b, dict):
            return float(b.get("total_hours", 0.0))
        try:
            return float(b or 0.0)
        except Exception:
            return 0.0

    combined[col_total] = combined["Matter Number"].map(_cycle_total)

    # ===== Per-user cycle hour columns (the part you asked for) =====
    # Gather every user observed in cycle_data
    users: List[str] = []
    for dat in cycle_data.values():
        if isinstance(dat, dict):
            for u in (dat.get("user_hours") or {}):
                users.append(u)
    users = sorted(set(users))

    # Add a column per user with the hours for that matter
    def _user_hours_for(dn: str, user: str) -> float:
        dat = cycle_data.get(dn, {})
        if not isinstance(dat, dict):
            return 0.0
        try:
            return float((dat.get("user_hours") or {}).get(user, 0.0) or 0.0)
        except Exception:
            return 0.0

    for user in users:
        col_user = f"{user} Cycle Hours ({cycle_start_label} - {cycle_end_label})"
        combined[col_user] = combined["Matter Number"].map(lambda dn, u=user: _user_hours_for(dn, u))

    # Expand custom fields onto columns
    for cf_name in CF_OUTPUT_FIELDS:
        combined[cf_name] = combined["_cf_map"].map(lambda d: (d or {}).get(cf_name, ""))

    # Select / order columns (match your sheet; include per-user columns right after total)
    # Select / order columns (match your sheet; Client Notes moves to after custom fields)
    base_cols = [
        "Matter Number","Client Name","CR ID","Net Trust Account Balance","Matter Stage",
        "Responsible Attorney","Main Paralegal","Supporting Attorney","Supporting Paralegal",
        # keep the TOTAL cycle hours next
        col_total,
    ]
    # then append the per-user cycle hour columns
    per_user_cols = [f"{u} Cycle Hours ({cycle_start_label} - {cycle_end_label})" for u in users]
    base_cols += per_user_cols
    # followed by the custom fields (in new order with renames and new empty columns)
    base_cols += [
        "Initial Client Goals","Initial Strategy","Has strategy changed Describe","Detailed List of Issues in the Case",
        "Strategy to Resolve Issues in the Case","Current action Items","Client Notes",
        "All Hearings on Calendar","All Deadlines on Calendar","DV situation description (DVRO: Attorney Action Required? Please Describe)",
        "Child Custody and Child Visitation (Timeshare): Temporary Orders Needed? State Current Order",
        "CS Add ons Extracurricular. Child Support: Temporary Orders Needed? State Current Order",
        "Spousal Support: Temporary Orders Needed? State Current Order",
        "Preliminary Declaration of Disclosures: Status of Client's PDDs and OPPs PDDs","Formal Discovery: Outline Discovery Strategy",
        "Parentage / Dissolution of Marriage / Legal Separation: Please Identify",
        "Judgment: Has a Judgement Been Entered? Please Specify","collection efforts",
        "Unbilled Hours"
    ]

    for c in base_cols:
        if c not in combined.columns:
            combined[c] = ""

    out = combined[base_cols].copy()
    out = out.sort_values(by="Net Trust Account Balance", ascending=False, kind="mergesort").reset_index(drop=True)
    return out

# =========================
# Excel formatting + save
# =========================
def apply_conditional_and_currency_formatting_with_totals(previous_cycle_df: pd.DataFrame,
                                                          mid_cycle_df: pd.DataFrame,
                                                          mid_cycle_data: Dict[str, Any],
                                                          current_cycle_data: Dict[str, Any],
                                                          mid_cycle_start_formatted: str,
                                                          mid_cycle_end_formatted: str,
                                                          current_date_formatted: str,
                                                          output_file: str) -> None:
    print(f"Applying formatting and saving to {output_file}...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for df in [previous_cycle_df, mid_cycle_df]:
            time_columns = [col for col in df.columns if 'Cycle Hours' in col]
            totals = df[time_columns].sum(numeric_only=True) if time_columns else pd.Series(dtype=float)
            totals_row = pd.Series('', index=df.columns)
            totals_row['Matter Number'] = 'TOTALS'
            for col in time_columns:
                totals_row[col] = totals.get(col, 0.0)
            df_with_totals = pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)
            sheet = 'Previous Billing Cycle' if df is previous_cycle_df else 'Mid Cycle'
            df_with_totals.to_excel(writer, sheet_name=sheet, index=False)

        # Totals sheet (user totals from cycle dicts)
        def user_totals(cycle_dict: Dict[str, Any]) -> Dict[str, float]:
            agg: Dict[str, float] = {}
            for v in cycle_dict.values():
                if not isinstance(v, dict):
                    continue
                for u, h in (v.get("user_hours") or {}).items():
                    try:
                        agg[u] = agg.get(u, 0.0) + float(h or 0.0)
                    except Exception:
                        pass
            return agg

        mid_cycle_totals = user_totals(mid_cycle_data)
        current_cycle_totals = user_totals(current_cycle_data)

        mid_totals_df = pd.DataFrame([
            {'User': u, f'Cycle Hours ({mid_cycle_start_formatted} - {mid_cycle_end_formatted})': h}
            for u, h in sorted(mid_cycle_totals.items(), key=lambda x: x[1], reverse=True)
        ])
        current_totals_df = pd.DataFrame([
            {'User': u, f'Cycle Running Total ({mid_cycle_start_formatted} - {current_date_formatted})': h}
            for u, h in sorted(current_cycle_totals.items(), key=lambda x: x[1], reverse=True)
        ])
        mid_totals_df.to_excel(writer, sheet_name='Billable Hour Totals', startrow=1, index=False)
        current_totals_df.to_excel(writer, sheet_name='Billable Hour Totals',
                                   startrow=mid_totals_df.shape[0] + 5, index=False)

    wb = load_workbook(output_file)
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    total_row_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    currency_style = NamedStyle(name='currency', number_format='$#,##0.00')
    bold_font = Font(bold=True)

    for sheet_name in ['Previous Billing Cycle', 'Mid Cycle']:
        ws = wb[sheet_name]
        last_row = ws.max_row
        net_balance_col = None
        time_cols: List[int] = []
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Net Trust Account Balance':
                net_balance_col = col_idx
            if 'Cycle Hours' in str(cell.value):
                time_cols.append(col_idx)

        # format rows except totals
        for row in ws.iter_rows(min_row=2, max_row=last_row-1):
            for col_idx, cell in enumerate(row, 1):
                if ws[1][col_idx - 1].value in ['Net Trust Account Balance']:
                    cell.style = currency_style
            if net_balance_col:
                net_cell = row[net_balance_col - 1]
                try:
                    val = float(net_cell.value or 0)
                    if val <= 0:
                        net_cell.fill = red_fill
                    elif 0 < val < 1000:
                        net_cell.fill = yellow_fill
                    else:
                        net_cell.fill = green_fill
                except Exception:
                    pass

        # totals row
        totals_row = ws[last_row]
        for cell in totals_row:
            cell.font = bold_font
            cell.fill = total_row_fill
            if cell.column in time_cols:
                cell.number_format = '#,##0.00'

        # add table excluding totals row
        table_ref = f"A1:{ws.cell(row=last_row-1, column=ws.max_column).coordinate}"
        table = Table(displayName=f"{sheet_name.replace(' ', '')}Table", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    wb.save(output_file)
    print(f" File saved: {output_file}")

# =========================
# SharePoint upload
# =========================
def ensure_folder(path: str, headers: Dict[str, str], site_id: str, drive_id: str) -> None:
    segs = path.strip("/").split("/")
    parent = ""
    for s in segs:
        full = f"{parent}/{s}" if parent else s
        url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root:/{full}"
        r = requests.get(url, headers=headers)
        if r.status_code == 404:
            if parent:
                create_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root:/{parent}:/children"
            else:
                create_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root/children"
            requests.post(create_url, headers=headers, json={
                "name": s, "folder": {}, "@microsoft.graph.conflictBehavior": "replace"
            }).raise_for_status()
        parent = full

def upload_to_sharepoint(file_path: str, file_name: str) -> None:
    TENANT_ID = os.getenv("SHAREPOINT_TENANT_ID") or os.getenv("GRAPH_TENANT_ID")
    APP_ID = os.getenv("SHAREPOINT_CLIENT_ID") or os.getenv("GRAPH_CLIENT_ID")
    APP_SECRET = os.getenv("SHAREPOINT_CLIENT_SECRET") or os.getenv("GRAPH_CLIENT_SECRET")
    SITE_ID = os.getenv("SHAREPOINT_SITE_ID")
    DRIVE_ID = os.getenv("SHAREPOINT_DRIVE_ID")
    LIBRARY_PATH = (os.getenv("SHAREPOINT_DOC_LIB") or "General Management/Global Case Review Lists").strip('"')

    authority = f"https://login.microsoftonline.com/{TENANT_ID}"
    app_conf = msal.ConfidentialClientApplication(APP_ID, authority=authority, client_credential=APP_SECRET)
    tok = app_conf.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in tok:
        raise Exception(f"Graph auth failed: {tok.get('error_description')}")
    headers = {"Authorization": f"Bearer {tok['access_token']}", "Content-Type": "application/json"}

    current_year = datetime.now().strftime("%Y")
    current_month = datetime.now().strftime("%m %B %Y")
    folder_path = f"{LIBRARY_PATH}/{current_year}/{current_month}"
    ensure_folder(folder_path, headers, SITE_ID, DRIVE_ID)

    encoded = quote(f"{folder_path}/{file_name}")
    upload_url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/drives/{DRIVE_ID}/root:/{encoded}:/content"

    with open(file_path, "rb") as f:
        res = requests.put(upload_url, headers={"Authorization": headers["Authorization"]}, data=f)
    if res.status_code not in (200, 201):
        raise Exception(f"Upload failed: {res.status_code} - {res.text[:200]}")
    print(f" Uploaded {file_name} to SharePoint at {folder_path}/")

# =========================
# Helpers
# =========================
def iso_range_from_labels(start_label_mmddyy: str, end_label_mmddyy: str, tz: str = "-08:00") -> Tuple[str, str]:
    # expects labels like "05/07/25"
    start = datetime.strptime(start_label_mmddyy, "%m/%d/%y").strftime("%Y-%m-%d")
    end = datetime.strptime(end_label_mmddyy, "%m/%d/%y").strftime("%Y-%m-%d")
    return f"{start}T00:00:00{tz}", f"{end}T23:59:59{tz}"

# =========================
# Orchestrator
# =========================
def fetch_and_process_data() -> None:
    # WINDOWS you care about
    prev_start_lbl, prev_end_lbl = "11/19/25", "12/02/25"
    mid_start_lbl,  mid_end_lbl  = "12/03/25", "12/16/25"
    tz = os.getenv("CLIO_TZ_OFFSET", "-08:00")

    prev_start_iso, prev_end_iso = iso_range_from_labels(prev_start_lbl, prev_end_lbl, tz)
    mid_start_iso, mid_end_iso   = iso_range_from_labels(mid_start_lbl, mid_end_lbl, tz)
    current_date = datetime.now()
    current_cycle_start_iso = mid_start_iso
    current_cycle_end_iso   = f"{current_date.strftime('%Y-%m-%d')}T23:59:59{tz}"

    # Build base frames once
    matter_df, ocb_df, bill_df = build_base_frames()

    # Cycle hours dicts (with user breakdown)
    previous_cycle_hours = fetch_billable_hours(prev_start_iso, prev_end_iso)
    mid_cycle_hours      = fetch_billable_hours(mid_start_iso, mid_end_iso)
    current_cycle_hours  = fetch_billable_hours(current_cycle_start_iso, current_cycle_end_iso)

    # Compose sheets (correct joins + per-user columns)
    previous_cycle_df = make_report_df(matter_df, ocb_df, bill_df, previous_cycle_hours, prev_start_lbl, prev_end_lbl)
    mid_cycle_df      = make_report_df(matter_df, ocb_df, bill_df, mid_cycle_hours,  mid_start_lbl,  mid_end_lbl)

    # Save & format
    current_date_str = datetime.now().strftime("%Y-%m-%d %I%p").lstrip('0').replace('.0', '.')
    output_file = f"TLR {current_date_str}.xlsx"
    apply_conditional_and_currency_formatting_with_totals(
        previous_cycle_df,
        mid_cycle_df,
        mid_cycle_hours,
        current_cycle_hours,
        mid_start_lbl,
        mid_end_lbl,
        current_date.strftime("%m/%d/%y"),
        output_file
    )

    print(f"\nUploading {output_file} to SharePoint...")
    upload_to_sharepoint(output_file, output_file)

    try:
        os.remove(output_file)
    except Exception:
        pass
    print(" Done.")

# =========================
# OAuth callback (optional local use)
# =========================
@app.route('/callback')
def callback():
    auth_code = request.args.get('code')
    if not auth_code:
        return "Missing auth code", 400
    resp = session.post(CLIO_TOKEN_URL, data={
        'grant_type': 'authorization_code',
        'code': auth_code,
        'client_id': CLIENT_ID,
        'client_secret': CLIENT_SECRET,
        'redirect_uri': REDIRECT_URI
    }, timeout=60)
    if resp.status_code != 200:
        return f"Authorization failed: {resp.status_code}", 400
    tokens = resp.json() or {}
    save_tokens_env(tokens)
    fetch_and_process_data()
    return 'Authorization complete. Data processing initiated.'

# =========================
# Entrypoint
# =========================
if __name__ == '__main__':
    try:
        _get_access_token()  # ensure we have an access token (refresh if needed)
        print("Access token ok. Starting…")
        fetch_and_process_data()
    except Exception as e:
        print(f"❌ Error: {e}")
